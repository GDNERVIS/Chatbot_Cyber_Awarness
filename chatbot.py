from time import sleep  # Importa a função sleep do módulo time para pausar a execução do programa
import os  # Interação com o sistema operacional
import sys  # Encaminhar argumentos para a aplicação
import gradio as gr  # Importa a biblioteca Gradio para criar interfaces de usuário
from openai import OpenAI  # Importa a classe OpenAI para interagir com a API OpenAI
from newsapi import NewsApiClient  # Importa a classe NewsApiClient para interagir com a API de notícias
import google.generativeai as genai
import json  # Importa a biblioteca json para manipulação de dados JSON
from llama_index.core.evaluation import SemanticSimilarityEvaluator
from llamaapi import LlamaAPI
import pandas as pd
import asyncio
from openpyxl import load_workbook

newsapi = NewsApiClient(api_key='YOUT API')  # Inicializa o cliente da API de notícias com a chave da API
client = OpenAI()  # Chama a API para conversar com o chat.
genai.configure(api_key="YOUR API") #API Google
llama = LlamaAPI("YOUR API") #chave Llma

def openai_chat(text,tituloseurls):
    # Função que envia uma mensagem para a API OpenAI e retorna a resposta
    text_to_chat = text  # Armazena o texto a ser enviado para a API
    titulos_e_urls = tituloseurls
    completion = client.chat.completions.create(
        model="gpt-4o",  # Especifica o modelo a ser usado
        messages=[  # Lista de mensagens que compõem a conversa
            {"role": "system", "content": "Você é um especialista em cybersecurity e possui mastria em se comunicar de forma facil sobre os assuntos da area de forma simples, clara e objetiva. Você provê conhecimento, comportamentos e atitudes. Você só pode responder as perguntas sobre cybersecurity, tudo que fugir do tema de segurança da informação você nao deve responder.\n Alem disso, o input que voce receberá contem: A pergunta e uma lista de tuplas. Cada tupla contém dois elementos: o título (article['title']) e a URL (article['url']) de um artigo.\n ou seja, voce deve responder sempre a pergunta e posteriormente fornecer o  titulo e o link(https://.....) de acesso da noticia que mais sentido com o contexto. Caso a tupla esteja vazia, não mencione nada."},  # Mensagem do sistema que define o contexto do assistente
            {"role": "user", "content": f"{text_to_chat} {titulos_e_urls}"}  # Mensagem do usuário que contém o texto a ser enviado
        ]
    )
    resposta = (completion.choices[0].message)  # Obtém a resposta da API
    resposta_limpa = resposta.content  # Extrai o conteúdo da resposta
    return resposta_limpa  # Retorna a resposta limpa

def gemini_chat(text,tituloseurls):
  text_to_chat = text  # Armazena o texto a ser enviado para a API
  titulos_e_urls = tituloseurls
  model = genai.GenerativeModel(
    model_name="gemini-1.5-flash",
    system_instruction="Você é um especialista em cybersecurity e possui mastria em se comunicar de forma facil sobre os assuntos da area de forma simples, clara e objetiva. Você provê conhecimento, comportamentos e atitudes. Você só pode responder as perguntas sobre cybersecurity, tudo que fugir do tema de segurança da informação você nao deve responder.\n Alem disso, o input que voce receberá contem: A pergunta e uma lista de tuplas. Cada tupla contém dois elementos: o título (article['title']) e a URL (article['url']) de um artigo.\n ou seja, voce deve responder sempre a pergunta e posteriormente fornecer o  titulo e o link(https://.....) de acesso da noticia que mais sentido com o contexto. Caso a tupla esteja vazia, não mencione nada."
    )
  resposta= model.generate_content(f"{text_to_chat} {titulos_e_urls}")
  resposta_limpa = resposta.text  # Extrai o conteúdo da resposta
  return resposta_limpa  # Retorna a resposta limpa

def llma_chat(openai, gemini ):
    output_openai = openai
    output_gemini =  gemini
    

    api_request_json = {
    "model": "llama3.1-405b",
    "max_tokens": 4046,
    "messages": [
        {"role": "system", "content": "Você é um especialista em cybersecurity e ira receber duas respostas oriundas outros modelos de ia generativos. Gere uma nova resposta com base nessas duas entradas usando as seguintes diretrzies:\nVocê é um especialista em cybersecurity e possui mastria em se comunicar de forma facil sobre os assuntos da area de forma simples, clara e objetiva. Você provê conhecimento, comportamentos e atitudes. Você só pode responder as perguntas sobre cybersecurity, tudo que fugir do tema de segurança da informação você nao deve responder. Você deve fornecer titulo e o link(https://.....) de acesso da noticia ao final da resposta."},
        {"role": "user", "content": f"Resposta OPENAI:{output_openai}\n\n RESPOSTA GEMINI {output_gemini}"},
        ]
    }

    response = llama.run(api_request_json)
    resposta_limpa = response.json().get('choices')[0].get('message').get('content')
    return resposta_limpa
    


def processa_to_news_API(text):
    text_to_chat = text
    completion = client.chat.completions.create(
        model="gpt-4o",  # Especifica o modelo a ser usado
        messages=[
            {"role": "system", "content": "Você precisa extrair da pergunta o tema. Abaixo segue o exemplo:\n Pergunta01: O que é phishing? Resposta de extração: Phishing Pergunta02: Como me proeteger de golpes digitais? Resposta: Golpes digitais"},  # Mensagem do sistema que define o contexto do assistente
            {"role": "user", "content": text_to_chat}  # Mensagem do usuário que contém o texto a ser enviado
        ]
    )
    resposta = (completion.choices[0].message)  # Obtém a resposta da API
    resposta_limpa = resposta.content  # Extrai o conteúdo da resposta
    return resposta_limpa

async def evaluate_similarity(response, reference):
    evaluator = SemanticSimilarityEvaluator()
    result = await evaluator.aevaluate(
        response=response,
        reference=reference,
    )
    return result



with gr.Blocks() as Demo:  # Cria um bloco de interface usando Gradio
    state = gr.State()  # Cria um estado para manter o histórico da conversa
    gr.Markdown("""<h1> Cyber Assistant</h1>""")  # Título da interface
    msg = gr.Textbox(placeholder="What do you have in mind?")  # Caixa de texto para entrada do usuário
    chatbot = gr.Chatbot()  # Componente de chatbot para exibir a conversa
    
    ints = [msg, state]  # Entradas para a função de resposta
    outs = [chatbot, state]  # Saídas da função de resposta

    def Respond_Chat(input, chat_history):
        # Função que processa a entrada do usuário e gera uma resposta
        chat_history = chat_history or []  # Inicializa o histórico de chat se estiver vazio
        sum_chat_history = list(sum(chat_history, ()))  # Achata o histórico de chat
        sum_chat_history.append(input)  # Adiciona a nova entrada ao histórico
        sum_chat_history_inp = ' '.join(sum_chat_history)  # Concatena o histórico em uma única string
        
        tema = processa_to_news_API(input) # retira o tema da pergunta
        noticias = newsapi.get_everything(q= tema) # seleciona as noticas com base no tema da pergunta
        titulos_e_urls = [(article['title'], article['url']) for article in noticias['articles']]#extrai as informações para um lista
        #print(titulos_e_urls)
        
        output_openai = openai_chat(input, titulos_e_urls)
        output_gemini = gemini_chat(input, titulos_e_urls)
        #print(output_openai)
        #print(output_gemini)
        output_llma = llma_chat(output_openai, output_gemini)
    
        #print(output_llma)
        result_similarity = asyncio.run(evaluate_similarity(output_openai, output_gemini)) #similaridade com as respostas orginais dos modelos da OpenAI e GPT-4o

        result_similarity_llma_vs_openai = asyncio.run(evaluate_similarity(output_llma, output_openai))
        result_similarity_llma_vs_gemini = asyncio.run(evaluate_similarity(output_llma, output_gemini))
       
        output = f"RESPOSTA OPENAI:\n\n{output_openai}\n\nRESPOSTA GEMINI:\n\n{output_gemini}\n\n RESULTADOS DA SIMILARIDADE ENTRE AS RESPOSTAS:\n\n Score:{result_similarity.score}\n\n Similaridade:{result_similarity.passing}\n\n RESPOSTA GERADA PELA SIMILARIDADE:{output_llma}"
        
        #print(output)#controle do output
        chat_history.append((input, output))  # Adiciona a entrada e a resposta ao histórico

        # Cria um DataFrame com os dados das respostas, separando por modelo
        dados = {
            "Pergunta":[input],
            "Resposta modelo gpt-4o": [output_openai],
            "Resposta modelo gemini-1.5-flash": [output_gemini],
            "Score": [result_similarity.score],  # Adiciona o score 
            "Similaridade": [result_similarity.passing],
            "Resposta com base dos modelos usados através do llama3.1-405b  ":[output_llma],
            "Score de similaridade com GPT-4o":[result_similarity_llma_vs_openai.score],
            "Similaridade com GPT-4o":[result_similarity_llma_vs_openai.passing],
            "Score de similaridade com gemini-1.5-flash":[result_similarity_llma_vs_gemini.score],
            "Similaridade com gemini-1.5-flash":[result_similarity_llma_vs_gemini.passing]
        }
        df = pd.DataFrame(dados)

        file_path = "resultados_similaridade.xlsx"
        sheet_name = "Sheet1"

        try:
            # Carregar o arquivo Excel existente
            workbook = load_workbook(file_path)
            # Verificar se a planilha já existe
            if sheet_name in workbook.sheetnames:
                # Abrir a planilha existente
                with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    # Encontrar o número de linhas existentes na planilha
                    existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                    startrow = len(existing_df) + 1  # Começar abaixo dos dados existentes
                    # Adicionar os novos dados
                    df.to_excel(writer, index=False, header=False, startrow=startrow, sheet_name=sheet_name)
            else:
                # Caso a planilha não exista, criar uma nova
                with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
        except FileNotFoundError:
                # Caso o arquivo não exista, criar um novo
                print("error")

        return chat_history, chat_history  # Retorna o histórico atualizado

    msg.submit(Respond_Chat, inputs=ints, outputs=outs)  # Envia a entrada do usuário para a função de resposta
    
Demo.launch(debug=True)  # Inicia a interface do Gradio em modo de depuração
